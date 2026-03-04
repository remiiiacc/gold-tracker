"""
Gold Tracker API Server
Flask backend — all API keys stay server-side, never in frontend.
Run with: gunicorn -w 2 -b 127.0.0.1:5000 app:app
"""
import io
import json
import os
import time
from datetime import datetime
from flask import Flask, jsonify, request
from flask_cors import CORS
from fetcher import fetch_fred, fetch_cot, fetch_yfinance
import pdfplumber

app = Flask(__name__)
CORS(app, origins=['http://gold.hb-labs.com', 'http://159.223.44.23'])

CACHE_DIR = '/opt/gold-tracker/api/cache'
os.makedirs(CACHE_DIR, exist_ok=True)

CACHE_TTL = {
    'fred.json': 24 * 3600,       # 24 hours
    'cot.json':  7 * 24 * 3600,   # 7 days
    'yfinance.json': 24 * 3600,   # 24 hours
}


def cache_path(filename):
    return os.path.join(CACHE_DIR, filename)


def is_fresh(filename):
    path = cache_path(filename)
    if not os.path.exists(path):
        return False
    age = time.time() - os.path.getmtime(path)
    return age < CACHE_TTL.get(filename, 24 * 3600)


def read_cache(filename):
    with open(cache_path(filename)) as f:
        return json.load(f)


def write_cache(filename, data):
    with open(cache_path(filename), 'w') as f:
        json.dump(data, f)
    return data


STATUS_JSON_PATH = '/var/www/gold-tracker/data/status.json'


def compute_analytics():
    """Compute all dashboard analytics from cached data. Returns a dict."""
    def load(name):
        p = cache_path(name)
        if not os.path.exists(p):
            return None, None
        with open(p) as f:
            return json.load(f), os.path.getmtime(p)

    fred, fred_mtime = load('fred.json')
    cot_data, cot_mtime = load('cot.json')
    yf, yf_mtime = load('yfinance.json')
    if not fred or not yf:
        return None

    # --- Monthly pairs for dual-regime regression ---
    tips_monthly = [t for t in (fred.get('tips_monthly') or []) if t.get('value') is not None]
    gold_monthly = fred.get('gold_monthly') or []
    gm_map = {g['date']: g['value'] for g in gold_monthly}
    pairs = sorted(
        [{'date': t['date'], 'tips': t['value'], 'gold': gm_map[t['date']]}
         for t in tips_monthly if t['date'] in gm_map],
        key=lambda x: x['date'],
    )
    latest = pairs[-1] if pairs else None
    regime1 = [(p['tips'], p['gold']) for p in pairs if p['date'] <= '2021-12-31']
    regime2 = [(p['tips'], p['gold']) for p in pairs if p['date'] >= '2022-01-01']

    def ols(pts):
        n = len(pts)
        if n < 2:
            return None
        sx = sum(x for x, y in pts); sy = sum(y for x, y in pts)
        sxy = sum(x * y for x, y in pts); sx2 = sum(x * x for x, y in pts)
        d = n * sx2 - sx * sx
        if not d:
            return None
        sl = (n * sxy - sx * sy) / d
        ic = (sy - sl * sx) / n
        ym = sy / n
        ss_tot = sum((y - ym) ** 2 for x, y in pts)
        ss_res = sum((y - (sl * x + ic)) ** 2 for x, y in pts)
        r2v = 1 - ss_res / ss_tot if ss_tot else 0
        return {'slope': round(sl, 2), 'intercept': round(ic, 2), 'r2': round(r2v, 4), 'n': n}

    r1 = ols(regime1)
    r2 = ols(regime2)
    current_tips = latest['tips'] if latest else None
    current_gold = latest['gold'] if latest else None
    r1_pred = round(r1['slope'] * current_tips + r1['intercept'], 2) if r1 and current_tips is not None else None
    r2_pred = round(r2['slope'] * current_tips + r2['intercept'], 2) if r2 and current_tips is not None else None
    premium = round(current_gold - r1_pred, 2) if current_gold is not None and r1_pred is not None else None
    premium_pct = round(premium / r1_pred * 100, 1) if premium is not None and r1_pred else None

    # --- Real rates ---
    def last_valid(series):
        arr = [x for x in (series if isinstance(series, list) else []) if x.get('value') is not None]
        return arr[-1] if arr else None

    tips_daily = sorted([x for x in (fred.get('tips') or []) if x.get('value') is not None], key=lambda x: x['date'])
    dxy_daily  = sorted([x for x in (fred.get('dxy')  or []) if x.get('value') is not None], key=lambda x: x['date'])
    sofr_rec   = last_valid(fred.get('sofr', []))
    tips_3m_chg = round(tips_daily[-1]['value'] - tips_daily[max(0, len(tips_daily) - 63)]['value'], 3) if len(tips_daily) >= 2 else None
    dxy_90d_chg = round(dxy_daily[-1]['value']  - dxy_daily[max(0, len(dxy_daily) - 90)]['value'],   3) if len(dxy_daily) >= 2 else None

    # --- COT ---
    cot_rows   = (cot_data or {}).get('data', [])
    cot_latest = cot_rows[-1] if cot_rows else {}
    cot_net    = cot_latest.get('net_long', 0)
    window     = sorted(r['net_long'] for r in cot_rows[-260:]) if cot_rows else [cot_net]
    cot_pctile = round(sum(1 for v in window if v <= cot_net) / len(window) * 100, 1) if window else None

    # --- Ratios ---
    def yf_close(key):
        arr = yf.get(key, [])
        return (arr[-1] if isinstance(arr, list) and arr else {}).get('close')

    gld_p = yf_close('gld'); gdx_p = yf_close('gdx')
    gf_p  = yf_close('gold_futures'); sf_p = yf_close('silver_futures')
    GLD_OZ = 0.09334
    spot   = gld_p / GLD_OZ if gld_p else None
    cobasis = round(spot - gf_p, 2) if spot is not None and gf_p else None
    gsr     = round(gf_p / sf_p, 2) if gf_p and sf_p else None
    gdxgld  = round(gdx_p / gld_p, 4) if gdx_p and gld_p else None
    gld_arr = yf.get('gld', []); gdx_arr = yf.get('gdx', [])
    gdxgld_3yavg = None
    if isinstance(gld_arr, list) and isinstance(gdx_arr, list) and len(gld_arr) >= 756:
        gm2 = {x['date']: x['close'] for x in gld_arr}
        r3y = [g['close'] / gm2[g['date']] for g in gdx_arr[-756:] if g['date'] in gm2]
        gdxgld_3yavg = round(sum(r3y) / len(r3y), 4) if r3y else None

    # --- Scorecard signals (9 signals, matching analytics.js buildScorecard) ---
    def sig(v, bull_fn, bear_fn):
        if v is None:
            return 'neutral'
        if bull_fn(v):
            return 'bullish'
        if bear_fn(v):
            return 'bearish'
        return 'neutral'

    gdxgld_vs_avg = round(gdxgld - gdxgld_3yavg, 4) if gdxgld and gdxgld_3yavg else None

    # ── WGC CB demand signal ────────────────────────────────────────────
    wgc_signal = 'neutral'
    wgc_value  = None
    wgc_label_detail = None

    if os.path.exists(WGC_DEMAND_JSON_PATH):
        try:
            with open(WGC_DEMAND_JSON_PATH) as f:
                wgc_rows = json.load(f)
            # Sort by quarter string ascending (YYYY-QN sorts correctly as string)
            wgc_rows.sort(key=lambda r: r.get('quarter', ''))
            cb_vals = [r['central_banks'] for r in wgc_rows
                       if r.get('central_banks') is not None]
            if len(cb_vals) >= 4:
                latest_q        = cb_vals[-1]
                trailing_4q_avg = sum(cb_vals[-4:]) / 4
                # YoY: compare latest quarter vs same quarter one year prior (4 back)
                yoy_pct = None
                if len(cb_vals) >= 5:
                    yoy_pct = ((cb_vals[-1] - cb_vals[-5]) / abs(cb_vals[-5]) * 100
                               if cb_vals[-5] != 0 else None)

                # Signal: bullish if latest > trailing 4Q avg AND yoy > 0
                if latest_q > trailing_4q_avg and (yoy_pct is None or yoy_pct > 0):
                    wgc_signal = 'bullish'
                elif latest_q < trailing_4q_avg * 0.8 and (yoy_pct is not None and yoy_pct < 0):
                    wgc_signal = 'bearish'
                else:
                    wgc_signal = 'neutral'

                latest_quarter_label = wgc_rows[-1].get('quarter', '')
                wgc_value = round(latest_q, 1)
                yoy_str = f'{yoy_pct:+.1f}% YoY' if yoy_pct is not None else 'N/A YoY'
                wgc_label_detail = f'{latest_quarter_label}: {wgc_value}t ({yoy_str})'
        except Exception:
            pass  # leave neutral on any error

    signals = {
        'realRateTrend':    {'value': tips_3m_chg,   'signal': sig(tips_3m_chg,   lambda v: v < -0.1, lambda v: v > 0.1),  'label': '3M TIPS change'},
        'rateModelPremium': {'value': premium_pct,   'signal': sig(premium_pct,   lambda v: v < 10,   lambda v: v > 20),   'label': 'Regime break premium %'},
        'dxyTrend':         {'value': dxy_90d_chg,   'signal': sig(dxy_90d_chg,   lambda v: v < -1,   lambda v: v > 1),    'label': '90d DXY change'},
        'cotPercentile':    {'value': cot_pctile,    'signal': sig(cot_pctile,    lambda v: v < 20,   lambda v: v > 80),   'label': '5Y COT percentile'},
        'cobasis':          {'value': cobasis,       'signal': sig(cobasis,       lambda v: v > 0,    lambda v: v < -2),   'label': 'Spot minus futures ($/oz)'},
        'gdxGldVs3yAvg':    {'value': gdxgld_vs_avg, 'signal': sig(gdxgld_vs_avg, lambda v: v > 0,    lambda v: v < -0.01),'label': 'GDX/GLD vs 3Y avg'},
        'goldSilverRatio':  {'value': gsr,           'signal': sig(gsr,           lambda v: v < 70,   lambda v: v > 80),   'label': 'GC=F / SI=F'},
        'cbDemand':         {'value': wgc_label_detail or wgc_value, 'signal': wgc_signal,            'label': 'Central bank demand (WGC)'},
        'etfFlowTrend':     {'value': None,           'signal': 'neutral',                                                   'label': 'ETF flow momentum'},
    }
    bullish_count = sum(1 for s in signals.values() if s['signal'] == 'bullish')

    def freshness(mtime):
        if not mtime:
            return None
        return {
            'lastUpdated': datetime.utcfromtimestamp(mtime).strftime('%Y-%m-%dT%H:%M:%SZ'),
            'ageHours': round((time.time() - mtime) / 3600, 1),
        }

    return {
        'lastUpdated':            datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
        'lastDate':               latest['date'] if latest else None,
        'currentGoldPrice':       current_gold,
        'currentTIPS':            current_tips,
        'regimeBreakPremium':     premium,
        'regimeBreakPremiumPct':  premium_pct,
        'regime1': {
            **(r1 or {}),
            'label':                   '2006-2021',
            'predictedAtCurrentRate':  r1_pred,
        },
        'regime2': {
            **(r2 or {}),
            'label':                   '2022-present',
            'predictedAtCurrentRate':  r2_pred,
        },
        'realRates': {
            'tips10y':      current_tips,
            'tips3mChange': tips_3m_chg,
            'dxy':          dxy_daily[-1]['value'] if dxy_daily else None,
            'dxy90dChange': dxy_90d_chg,
            'sofr':         sofr_rec['value'] if sofr_rec else None,
        },
        'cot': {
            'netLong':       cot_net,
            'netLongPct':    cot_latest.get('net_long_pct'),
            'openInterest':  cot_latest.get('open_interest'),
            'percentile5y':  cot_pctile,
            'date':          cot_latest.get('date'),
        },
        'ratios': {
            'goldSilverRatio': gsr,
            'cobasis':         cobasis,
            'gdxGld':          gdxgld,
            'gdxGld3yAvg':     gdxgld_3yavg,
        },
        'scorecard': {
            'bullishSignals': bullish_count,
            'totalSignals':   len(signals),
            'composite':      f'{bullish_count}/{len(signals)}',
            'signals':        signals,
        },
        'dataFreshness': {
            'fred':     freshness(fred_mtime),
            'cot':      freshness(cot_mtime),
            'yfinance': freshness(yf_mtime),
        },
    }


def write_status_json():
    """Compute analytics and write to /var/www/gold-tracker/data/status.json."""
    try:
        data = compute_analytics()
        if data is None:
            return
        os.makedirs(os.path.dirname(STATUS_JSON_PATH), exist_ok=True)
        with open(STATUS_JSON_PATH, 'w') as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass  # Never let this break a data request


def serve(filename, fetcher_fn, force=False):
    """Generic cache-or-fetch handler."""
    if not force and is_fresh(filename):
        try:
            data = read_cache(filename)
            data['_cache'] = 'hit'
            return jsonify(data)
        except Exception:
            pass
    try:
        data = fetcher_fn()
        write_cache(filename, data)
        data['_cache'] = 'miss'
        write_status_json()  # Regenerate status.json after every fresh fetch
        return jsonify(data)
    except Exception as e:
        # Return cached data even if stale, rather than error
        if os.path.exists(cache_path(filename)):
            try:
                data = read_cache(filename)
                data['_cache'] = 'stale'
                data['_error'] = str(e)
                return jsonify(data)
            except Exception:
                pass
        return jsonify({'error': str(e), 'fetched_at': None}), 503


@app.route('/api/fred')
def fred_data():
    force = request.args.get('refresh') == '1'
    return serve('fred.json', fetch_fred, force=force)


@app.route('/api/cot')
def cot_data():
    force = request.args.get('refresh') == '1'
    return serve('cot.json', fetch_cot, force=force)


@app.route('/api/yfinance')
def yfinance_data():
    force = request.args.get('refresh') == '1'
    return serve('yfinance.json', fetch_yfinance, force=force)


@app.route('/api/status')
def status():
    files = {
        'fred': 'fred.json',
        'cot': 'cot.json',
        'yfinance': 'yfinance.json',
    }
    out = {}
    for key, fname in files.items():
        path = cache_path(fname)
        if os.path.exists(path):
            mtime = os.path.getmtime(path)
            out[key] = {
                'last_updated': datetime.utcfromtimestamp(mtime).isoformat() + 'Z',
                'fresh': is_fresh(fname),
                'age_hours': round((time.time() - mtime) / 3600, 1),
            }
        else:
            out[key] = {'last_updated': None, 'fresh': False, 'age_hours': None}
    return jsonify(out)


@app.route('/api/goldprice')
def gold_price():
    """Lightweight endpoint for the live gold price header stat.
    Reads from yfinance cache — fast, no external call needed."""
    try:
        yf_path = cache_path('yfinance.json')
        if not os.path.exists(yf_path):
            return jsonify({'error': 'No cache yet'}), 503

        with open(yf_path) as f:
            yf = json.load(f)

        futures = yf.get('gold_futures', [])
        if not isinstance(futures, list) or len(futures) < 2:
            return jsonify({'error': 'Insufficient data'}), 503

        current  = futures[-1]
        prev_day = futures[-2]
        price    = current['close']
        prev     = prev_day['close']
        day_chg  = price - prev
        day_pct  = day_chg / prev * 100

        # YoY: find ~252 trading days ago (1 year)
        yoy_idx   = max(0, len(futures) - 252)
        yoy_price = futures[yoy_idx]['close']
        yoy_pct   = (price - yoy_price) / yoy_price * 100

        return jsonify({
            'price':     round(price, 2),
            'day_chg':   round(day_chg, 2),
            'day_pct':   round(day_pct, 2),
            'yoy_pct':   round(yoy_pct, 2),
            'date':      current['date'],
            'as_of':     datetime.utcfromtimestamp(os.path.getmtime(yf_path)).isoformat() + 'Z',
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 503


@app.route('/api/analytics')
def analytics_export():
    """Return the latest computed analytics as JSON.
    Also regenerates /data/status.json. Useful for forcing a refresh."""
    data = compute_analytics()
    if data is None:
        return jsonify({'error': 'Data not yet cached'}), 503
    write_status_json()
    return jsonify(data)


@app.route('/api/health')
def health():
    return jsonify({'status': 'ok', 'time': datetime.utcnow().isoformat() + 'Z'})


SNAPSHOT_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Gold Tracker — Market Data Snapshot</title>
<style>
  body {{ font-family: monospace; background: #0f1419; color: #e6edf3; padding: 24px; max-width: 800px; margin: 0 auto; }}
  h1   {{ color: #d29922; margin-bottom: 4px; }}
  h2   {{ color: #58a6ff; margin-top: 24px; border-bottom: 1px solid #2d3a4a; padding-bottom: 6px; }}
  table{{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
  td,th{{ padding: 8px 12px; border-bottom: 1px solid #2d3a4a; text-align: left; }}
  th   {{ color: #6e7681; font-size: 12px; text-transform: uppercase; }}
  .bull{{ color: #3fb950; }} .bear{{ color: #f85149; }} .neut{{ color: #8b949e; }}
  .price{{ font-size: 28px; font-weight: bold; color: #d29922; }}
  .meta {{ color: #6e7681; font-size: 12px; }}
</style>
</head>
<body>
<h1>Gold Tracker — Market Snapshot</h1>
<p class="meta">Generated: {generated} | Data sources: FRED, CFTC, Yahoo Finance</p>
<p class="meta"><a href="/" style="color:#58a6ff;">← Back to dashboard</a></p>

<h2>Gold Price</h2>
<table>
<tr><th>Metric</th><th>Value</th><th>As of</th></tr>
<tr><td>Spot / Front-month Futures (GC=F)</td><td class="price">${gold_price:,.0f}/oz</td><td>{gold_date}</td></tr>
<tr><td>Day Change</td><td class="{day_cls}">{day_sign}{day_chg:,.0f} ({day_sign}{day_pct:.1f}%)</td><td>vs prev close</td></tr>
<tr><td>YoY Change</td><td class="{yoy_cls}">{yoy_sign}{yoy_pct:.1f}%</td><td>~252 trading days</td></tr>
</table>

<h2>Real Rates (FRED)</h2>
<table>
<tr><th>Series</th><th>Value</th><th>Date</th></tr>
<tr><td>10Y TIPS Yield (Real Rate)</td><td>{tips_val:.2f}%</td><td>{tips_date}</td></tr>
<tr><td>USD Broad Trade-Weighted Index (DXY)</td><td>{dxy_val:.2f}</td><td>{dxy_date}</td></tr>
<tr><td>SOFR</td><td>{sofr_val:.2f}%</td><td>{sofr_date}</td></tr>
<tr><td>Rate Model Implied Gold Price</td><td>${model_price:,.0f}/oz</td><td>OLS: gold = {slope:.0f} × TIPS + {intercept:.0f}</td></tr>
<tr><td>Speculative Premium vs Model</td><td class="{prem_cls}">{prem_sign}{premium:.1f}%</td><td>actual vs model-implied</td></tr>
</table>

<h2>CFTC Positioning (Non-Commercial / Speculative)</h2>
<table>
<tr><th>Metric</th><th>Value</th><th>Date</th></tr>
<tr><td>Net Long Contracts</td><td>{cot_net:,}</td><td>{cot_date}</td></tr>
<tr><td>Net Long % of Open Interest</td><td>{cot_pct:.1f}%</td><td></td></tr>
<tr><td>Positioning Percentile (5-year history)</td><td class="{cot_cls}">{cot_percentile:.0f}th percentile</td><td></td></tr>
<tr><td>Open Interest</td><td>{cot_oi:,}</td><td></td></tr>
</table>

<h2>Market Ratios</h2>
<table>
<tr><th>Metric</th><th>Value</th><th>Signal</th></tr>
<tr><td>Gold / Silver Ratio (GC=F ÷ SI=F)</td><td>{gsr:.1f}</td><td class="{gsr_cls}">{gsr_signal}</td></tr>
<tr><td>GDX / GLD Ratio (Miners vs Physical)</td><td>{gdxgld:.4f}</td><td class="{gdxgld_cls}">{gdxgld_signal}</td></tr>
<tr><td>Gold Cobasis (Spot − Futures, $/oz)</td><td>{cobasis:+.2f}</td><td class="{cb_cls}">{cb_state}</td></tr>
</table>

<h2>Bull Market Scorecard</h2>
<p><strong>Composite: <span class="{score_cls}">{bullish}/9 Bullish Signals</span></strong></p>
<table>
<tr><th>#</th><th>Signal</th><th>Value</th><th>Status</th></tr>
{scorecard_rows}
</table>

<h2>Data Freshness</h2>
<table>
<tr><th>Source</th><th>Last Updated (UTC)</th><th>Status</th></tr>
{freshness_rows}
</table>
</body>
</html>"""


def build_snapshot():
    """Compute all signals from cached data and return rendered HTML."""
    import json, os, time
    from datetime import datetime

    CACHE_DIR = '/opt/gold-tracker/api/cache'

    def load(name):
        p = os.path.join(CACHE_DIR, name)
        if not os.path.exists(p):
            return None, None
        with open(p) as f:
            return json.load(f), os.path.getmtime(p)

    fred, fred_mtime = load('fred.json')
    cot_data, cot_mtime   = load('cot.json')
    yf, yf_mtime   = load('yfinance.json')

    if not fred or not yf:
        return "<p>Data not yet cached. Try again shortly.</p>"

    # ---- Gold price ----
    gold_series = fred.get('gold', [])
    gf = sorted([x for x in (gold_series if isinstance(gold_series, list) else []) if x.get('value')], key=lambda x: x['date'])
    gold_price = gf[-1]['value'] if gf else 0
    gold_date  = gf[-1]['date']  if gf else '—'
    prev_price = gf[-2]['value'] if len(gf) >= 2 else gold_price
    day_chg = gold_price - prev_price
    day_pct = day_chg / prev_price * 100 if prev_price else 0
    yoy_price = gf[max(0, len(gf)-252)]['value'] if gf else gold_price
    yoy_pct = (gold_price - yoy_price) / yoy_price * 100 if yoy_price else 0

    # ---- TIPS, DXY, SOFR ----
    def last_valid(series):
        arr = [x for x in (series if isinstance(series, list) else []) if x.get('value') is not None]
        return arr[-1] if arr else {'value': 0, 'date': '—'}

    tips = last_valid(fred.get('tips', []))
    dxy  = last_valid(fred.get('dxy',  []))
    sofr = last_valid(fred.get('sofr', []))

    # ---- OLS regression: gold = slope * TIPS + intercept ----
    gold_map = {x['date']: x['value'] for x in gf}
    tips_series = [x for x in (fred.get('tips', []) or []) if x.get('value') is not None]
    pairs = [(t['value'], gold_map[t['date']]) for t in tips_series if t['date'] in gold_map]
    slope = intercept = r2 = 0
    model_price = premium = 0
    if len(pairs) >= 2:
        xs = [p[0] for p in pairs]; ys = [p[1] for p in pairs]
        n = len(xs); sx = sum(xs); sy = sum(ys)
        sxy = sum(x*y for x,y in zip(xs,ys)); sxx = sum(x*x for x in xs)
        denom = n*sxx - sx*sx
        if denom:
            slope = (n*sxy - sx*sy) / denom
            intercept = (sy - slope*sx) / n
        model_price = slope * tips['value'] + intercept
        premium = (gold_price - model_price) / model_price * 100 if model_price else 0

    # ---- COT ----
    cot_rows = (cot_data or {}).get('data', [])
    cot_latest = cot_rows[-1] if cot_rows else {}
    cot_net  = cot_latest.get('net_long', 0)
    cot_pct_oi = cot_latest.get('net_long_pct', 0)
    cot_oi   = cot_latest.get('open_interest', 0)
    cot_date_val = cot_latest.get('date', '—')
    # Percentile vs 5-year history
    five_yr_ago = cot_rows[max(0, len(cot_rows)-260)]['net_long'] if cot_rows else 0
    window = [r['net_long'] for r in cot_rows[-260:]] if cot_rows else [cot_net]
    window_sorted = sorted(window)
    cot_percentile = sum(1 for v in window_sorted if v <= cot_net) / len(window_sorted) * 100

    # ---- Ratios ----
    def yf_close(key):
        arr = yf.get(key, [])
        return (arr[-1] if isinstance(arr, list) and arr else {}).get('close', 0)

    gld_p  = yf_close('gld');  gdx_p = yf_close('gdx')
    gf_p   = yf_close('gold_futures'); sf_p = yf_close('silver_futures')
    GLD_OZ = 0.09334
    spot_price = gld_p / GLD_OZ if gld_p else 0
    cobasis = spot_price - gf_p if spot_price and gf_p else 0
    gsr = gf_p / sf_p if sf_p else 0
    gdxgld = gdx_p / gld_p if gld_p else 0

    # 3Y average GDX/GLD
    gld_arr = yf.get('gld', []); gdx_arr = yf.get('gdx', [])
    if isinstance(gld_arr, list) and isinstance(gdx_arr, list) and len(gld_arr) >= 756:
        gld_map_yf = {x['date']: x['close'] for x in gld_arr}
        ratios_3y = [g['close']/gld_map_yf[g['date']] for g in gdx_arr[-756:] if g['date'] in gld_map_yf]
        gdxgld_3yavg = sum(ratios_3y)/len(ratios_3y) if ratios_3y else gdxgld
    else:
        gdxgld_3yavg = gdxgld

    # ---- Signals ----
    # TIPS 3-month change
    tips_series_sorted = sorted(tips_series, key=lambda x: x['date'])
    tips_3m = tips_series_sorted[-1]['value'] - tips_series_sorted[max(0,len(tips_series_sorted)-63)]['value'] if tips_series_sorted else 0
    # DXY 90-day change
    dxy_sorted = sorted([x for x in (fred.get('dxy',[]) or []) if x.get('value') is not None], key=lambda x: x['date'])
    dxy_90d = dxy_sorted[-1]['value'] - dxy_sorted[max(0,len(dxy_sorted)-90)]['value'] if dxy_sorted else 0

    def sig(v, bull_test, bear_test, bull_label='Bullish', bear_label='Cautionary', neut_label='Neutral'):
        if bull_test(v): return 'bull', bull_label
        if bear_test(v): return 'bear', bear_label
        return 'neut', neut_label

    # ── WGC CB demand signal ────────────────────────────────────────────
    wgc_signal = 'neutral'
    wgc_value  = None
    wgc_label_detail = None

    if os.path.exists(QUARTERLY_JSON_PATH):
        try:
            with open(QUARTERLY_JSON_PATH) as f:
                wgc_rows = json.load(f)
            wgc_rows.sort(key=lambda r: r.get('quarter', ''))
            cb_vals = [r['central_banks'] for r in wgc_rows
                       if r.get('central_banks') is not None]
            if len(cb_vals) >= 4:
                latest_q        = cb_vals[-1]
                trailing_4q_avg = sum(cb_vals[-4:]) / 4
                yoy_pct = None
                if len(cb_vals) >= 5:
                    yoy_pct = ((cb_vals[-1] - cb_vals[-5]) / abs(cb_vals[-5]) * 100
                               if cb_vals[-5] != 0 else None)
                if latest_q > trailing_4q_avg and (yoy_pct is None or yoy_pct > 0):
                    wgc_signal = 'bullish'
                elif latest_q < trailing_4q_avg * 0.8 and (yoy_pct is not None and yoy_pct < 0):
                    wgc_signal = 'bearish'
                else:
                    wgc_signal = 'neutral'
                latest_quarter_label = wgc_rows[-1].get('quarter', '')
                wgc_value = round(latest_q, 1)
                yoy_str = f'{yoy_pct:+.1f}% YoY' if yoy_pct is not None else 'N/A YoY'
                wgc_label_detail = f'{latest_quarter_label}: {wgc_value}t ({yoy_str})'
        except Exception:
            pass  # leave neutral on any error

    signals = [
        ('Real Rate Trend (3M TIPS Δ)',      f'{tips_3m:+.2f}%',             *sig(tips_3m,      lambda v: v < -0.1,  lambda v: v > 0.1)),
        ('Rate Model Premium',               f'{premium:+.1f}%',              *sig(premium,      lambda v: v < 10,    lambda v: v > 20)),
        ('DXY Trend (90d)',                  f'{dxy_90d:+.2f}',               *sig(dxy_90d,      lambda v: v < -1,    lambda v: v > 1)),
        ('COT Positioning Percentile',       f'{cot_percentile:.0f}th pct',   *sig(cot_percentile, lambda v: v < 20, lambda v: v > 80)),
        ('Gold Cobasis',                     f'${cobasis:+.2f}/oz',           *sig(cobasis,      lambda v: v > 0,     lambda v: v < -2, 'Backwardation', 'Deep Contango')),
        ('Miner Confirmation (GDX/GLD)',     f'{gdxgld:.4f}',                 *sig(gdxgld - gdxgld_3yavg, lambda v: v > 0, lambda v: v < -0.01, 'Above 3Y avg', 'Below 3Y avg')),
        ('Gold/Silver Ratio',                f'{gsr:.1f}',                    *sig(gsr,          lambda v: v < 70,    lambda v: v > 80, 'Below 70 (normal)', 'Above 80 (late-cycle)')),
        ('CB Demand (latest quarter)',
         wgc_label_detail or 'No WGC data',
         'bull' if wgc_signal == 'bullish' else ('bear' if wgc_signal == 'bearish' else 'neut'),
         'WGC Gold Demand Trends'),
        ('ETF Flow Trend',                   '—',                             'neut', 'See dashboard'),
    ]

    bullish = sum(1 for s in signals if s[2] == 'bull')
    score_cls = 'bull' if bullish >= 6 else 'bear' if bullish <= 3 else 'neut'

    scorecard_rows = '\n'.join(
        f'<tr><td>{i+1}</td><td>{s[0]}</td><td>{s[1]}</td><td class="{s[2]}">{s[3]}</td></tr>'
        for i, s in enumerate(signals)
    )

    def fresh_row(label, mtime):
        if not mtime: return f'<tr><td>{label}</td><td>—</td><td class="bear">No cache</td></tr>'
        dt = datetime.utcfromtimestamp(mtime).strftime('%Y-%m-%d %H:%M')
        age_h = (time.time() - mtime) / 3600
        cls = 'bull' if age_h < 25 else 'neut' if age_h < 170 else 'bear'
        return f'<tr><td>{label}</td><td>{dt}</td><td class="{cls}">{age_h:.1f}h ago</td></tr>'

    freshness_rows = '\n'.join([
        fresh_row('FRED (TIPS, Gold, DXY, SOFR)', fred_mtime),
        fresh_row('CFTC COT (Speculative Positioning)', cot_mtime),
        fresh_row('Yahoo Finance (GDX, GLD, SLV, GC=F)', yf_mtime),
    ])

    return SNAPSHOT_TEMPLATE.format(
        generated=datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC'),
        gold_price=gold_price, gold_date=gold_date,
        day_sign='+' if day_chg >= 0 else '', day_chg=day_chg, day_pct=day_pct,
        day_cls='bull' if day_chg >= 0 else 'bear',
        yoy_sign='+' if yoy_pct >= 0 else '', yoy_pct=yoy_pct,
        yoy_cls='bull' if yoy_pct >= 0 else 'bear',
        tips_val=tips['value'], tips_date=tips['date'],
        dxy_val=dxy['value'],   dxy_date=dxy['date'],
        sofr_val=sofr['value'], sofr_date=sofr.get('date','—'),
        slope=slope, intercept=intercept, model_price=model_price,
        prem_sign='+' if premium >= 0 else '', premium=premium,
        prem_cls='bear' if premium > 20 else 'bull' if premium < 10 else 'neut',
        cot_net=cot_net, cot_pct=cot_pct_oi, cot_oi=cot_oi, cot_date=cot_date_val,
        cot_percentile=cot_percentile,
        cot_cls='bear' if cot_percentile > 80 else 'bull' if cot_percentile < 20 else 'neut',
        gsr=gsr, gsr_cls='bear' if gsr > 80 else 'bull' if gsr < 70 else 'neut',
        gsr_signal='Late-cycle (>80)' if gsr > 80 else 'Normal (<70)' if gsr < 70 else 'Elevated (70-80)',
        gdxgld=gdxgld, gdxgld_3yavg=gdxgld_3yavg,
        gdxgld_cls='bull' if gdxgld > gdxgld_3yavg else 'bear',
        gdxgld_signal=f'Above 3Y avg ({gdxgld_3yavg:.4f})' if gdxgld > gdxgld_3yavg else f'Below 3Y avg ({gdxgld_3yavg:.4f})',
        cobasis=cobasis,
        cb_cls='bull' if cobasis > 0 else 'bear' if cobasis < -2 else 'neut',
        cb_state='Backwardation (bullish)' if cobasis > 0 else 'Deep Contango' if cobasis < -2 else 'Contango',
        bullish=bullish, score_cls=score_cls,
        scorecard_rows=scorecard_rows,
        freshness_rows=freshness_rows,
    )


@app.route('/snapshot')
def snapshot():
    """Plain-HTML snapshot readable by AI crawlers, curl, screen readers."""
    from flask import Response
    html = build_snapshot()
    return Response(html, mimetype='text/html')



import threading

# ── /api/analysis — in-memory cache + rate limiting ────────────────────────
_analysis_cache = {}   # "{chart}_{YYYY-MM-DD}" -> {analysis, generated_at, stored_at, expires_at}
_rate_data      = {}   # ip -> {min_count, min_ts, hr_count, hr_ts}
_rate_lock      = threading.Lock()
ANALYSIS_TTL    = 3600  # 1-hour cache

VALID_CHARTS = {'chart1', 'chart2', 'chart3', 'chart4', 'chart5', 'chart6', 'chart7'}

CHART_PROMPTS = {
    'chart1': (
        "You are a concise gold market analyst. Analyse the dual-regime regression data below.\n\n"
        "Regime 1 (2006–2021): real TIPS yields were the dominant driver of gold (R² ≈ 0.88). "
        "Regime 2 (2022–present): the Fed raised real rates to 15-year highs yet gold surged — "
        "the historical relationship broke down.\n\n"
        "DATA:\n{context_json}\n\n"
        "Write exactly 3 paragraphs (no headers, no bullet points):\n"
        "1. What the regime-break premium magnitude tells us today about who is driving gold.\n"
        "2. Whether the premium signals overextension, structural repricing, or something in between.\n"
        "3. The single most important risk that could cause regime reversion, and what would instead sustain it.\n\n"
        "Be direct, cite specific numbers from the data. No financial advice. ≤350 words."
    ),
    'chart2': (
        "You are a concise gold market analyst. Analyse the US dollar (DXY) vs gold relationship from the data below.\n\n"
        "Historically gold and the dollar move inversely. Sustained deviations from this pattern are analytically significant.\n\n"
        "DATA:\n{context_json}\n\n"
        "Write exactly 3 paragraphs (no headers, no bullet points):\n"
        "1. What the 90-day DXY trend implies for near-term gold price pressure.\n"
        "2. Whether gold is tracking or diverging from the historical DXY relationship, and what that divergence (if any) means.\n"
        "3. The DXY level or catalyst that would most materially alter the gold outlook.\n\n"
        "Be direct, cite specific numbers from the data. No financial advice. ≤350 words."
    ),
    'chart3': (
        "You are a concise gold market analyst. Analyse the CFTC Commitments of Traders (COT) speculative positioning data below.\n\n"
        "Non-commercial (managed money) positioning in gold futures is a classic contrarian signal. "
        "Extreme long crowding often precedes pullbacks; extreme short positioning often precedes rallies.\n\n"
        "DATA:\n{context_json}\n\n"
        "Write exactly 3 paragraphs (no headers, no bullet points):\n"
        "1. What the net long level and 5-year percentile tell us about speculative sentiment right now.\n"
        "2. Whether the current positioning is a contrarian red flag, a sign of under-positioning, or neutral.\n"
        "3. What a significant positioning change in either direction would imply for price action.\n\n"
        "Be direct, cite specific numbers from the data. No financial advice. ≤350 words."
    ),
    'chart4': (
        "You are a concise gold market analyst. Analyse the gold cobasis (spot minus nearby futures) data below.\n\n"
        "A positive cobasis (backwardation) means immediate delivery trades at a premium to futures — "
        "a signal of physical scarcity and genuine demand. Deep contango suggests futures-driven speculation or abundant supply.\n\n"
        "DATA:\n{context_json}\n\n"
        "Write exactly 3 paragraphs (no headers, no bullet points):\n"
        "1. What the current cobasis reading says about the balance between physical and paper gold demand.\n"
        "2. Historical context for this reading and what cobasis regimes have tended to precede.\n"
        "3. How to weight this signal alongside the macro environment (rates, dollar, central bank demand).\n\n"
        "Be direct, cite specific numbers from the data. No financial advice. ≤350 words."
    ),
    'chart5': (
        "You are a concise gold market analyst. Analyse the gold/silver ratio data below.\n\n"
        "The ratio measures how many ounces of silver buy one ounce of gold. "
        "Below 70 is historically normal; above 80 signals late-cycle stress, risk-off conditions, or industrial demand weakness. "
        "Silver typically outperforms gold in strong commodity bull markets.\n\n"
        "DATA:\n{context_json}\n\n"
        "Write exactly 3 paragraphs (no headers, no bullet points):\n"
        "1. What the current ratio level implies about the relative pricing of the two metals.\n"
        "2. What the ratio says about the stage and character of the current gold cycle.\n"
        "3. Whether silver's relative position represents a structural lag, an opportunity, or a warning signal.\n\n"
        "Be direct, cite specific numbers from the data. No financial advice. ≤350 words."
    ),
    'chart6': (
        "You are a concise gold market analyst. Analyse the GDX/GLD ratio (gold miners vs physical gold ETF) data below.\n\n"
        "When miners outperform physical gold (ratio rising above the 3-year average), it signals institutional conviction — "
        "miners provide leveraged exposure and sophisticated investors tend to rotate into them only when they believe "
        "the gold move is durable. Miner underperformance during a gold rally is a divergence worth noting.\n\n"
        "DATA:\n{context_json}\n\n"
        "Write exactly 3 paragraphs (no headers, no bullet points):\n"
        "1. What the GDX/GLD ratio vs the 3-year average tells us about institutional conviction in the gold move.\n"
        "2. Whether miner performance is confirming or diverging from gold price, and what that divergence implies.\n"
        "3. The key operational or valuation factors that could drive a re-rating of miners vs physical.\n\n"
        "Be direct, cite specific numbers from the data. No financial advice. ≤350 words."
    ),
    'chart7': (
        "You are a concise gold market analyst. Analyse the composite bull market scorecard below.\n\n"
        "The scorecard tracks 7 quantifiable signals across real rates, dollar strength, speculative positioning, "
        "physical demand (cobasis), miner confirmation, and the gold/silver ratio. "
        "Two signals — central bank demand and ETF flow trend — are qualitative and currently held neutral.\n\n"
        "DATA:\n{context_json}\n\n"
        "Write exactly 3 paragraphs (no headers, no bullet points):\n"
        "1. The overall composite reading and which 2–3 individual signals are most analytically meaningful right now.\n"
        "2. The key tension or contradiction in the signals (e.g., bearish positioning vs bullish price, or rate headwinds vs CB demand).\n"
        "3. Which signal change would most shift the composite score, and what event or data point would trigger it.\n\n"
        "Be direct, cite specific numbers from the data. No financial advice. ≤350 words."
    ),
}

CHART_CONTEXT_KEYS = {
    'chart1': lambda s: {
        'currentGoldPrice':      s.get('currentGoldPrice'),
        'currentTIPS':           s.get('currentTIPS'),
        'regimeBreakPremium':    s.get('regimeBreakPremium'),
        'regimeBreakPremiumPct': s.get('regimeBreakPremiumPct'),
        'regime1':               s.get('regime1'),
        'regime2':               s.get('regime2'),
        'lastDate':              s.get('lastDate'),
    },
    'chart2': lambda s: {
        'currentGoldPrice': s.get('currentGoldPrice'),
        'realRates': s.get('realRates'),
        'scorecard_dxy': (s.get('scorecard') or {}).get('signals', {}).get('dxyTrend'),
    },
    'chart3': lambda s: {
        'cot': s.get('cot'),
        'scorecard_cot': (s.get('scorecard') or {}).get('signals', {}).get('cotPercentile'),
    },
    'chart4': lambda s: {
        'cobasis':           (s.get('ratios') or {}).get('cobasis'),
        'currentGoldPrice':  s.get('currentGoldPrice'),
        'scorecard_cobasis': (s.get('scorecard') or {}).get('signals', {}).get('cobasis'),
    },
    'chart5': lambda s: {
        'goldSilverRatio':    (s.get('ratios') or {}).get('goldSilverRatio'),
        'scorecard_gsr':      (s.get('scorecard') or {}).get('signals', {}).get('goldSilverRatio'),
    },
    'chart6': lambda s: {
        'gdxGld':        (s.get('ratios') or {}).get('gdxGld'),
        'gdxGld3yAvg':   (s.get('ratios') or {}).get('gdxGld3yAvg'),
        'scorecard_gdx': (s.get('scorecard') or {}).get('signals', {}).get('gdxGldVs3yAvg'),
    },
    'chart7': lambda s: {
        'scorecard':        s.get('scorecard'),
        'currentGoldPrice': s.get('currentGoldPrice'),
        'currentTIPS':      s.get('currentTIPS'),
        'realRates':        s.get('realRates'),
    },
}


def _check_rate_limit(ip):
    """Returns True if the request is allowed, False if rate-limited."""
    now = time.time()
    with _rate_lock:
        d = _rate_data.get(ip, {'min_count': 0, 'min_ts': now, 'hr_count': 0, 'hr_ts': now})
        if now - d['min_ts'] >= 60:
            d['min_count'] = 0
            d['min_ts'] = now
        if now - d['hr_ts'] >= 3600:
            d['hr_count'] = 0
            d['hr_ts'] = now
        if d['min_count'] >= 10 or d['hr_count'] >= 50:
            return False
        d['min_count'] += 1
        d['hr_count'] += 1
        _rate_data[ip] = d
        return True


@app.route('/api/analysis', methods=['POST'])
def analysis():
    """Generate AI analysis for a given chart using cached market data."""
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()
    if not _check_rate_limit(ip):
        return jsonify({'error': 'Rate limit exceeded. Please wait before requesting again.'}), 429

    body = request.get_json(silent=True) or {}
    chart_id = str(body.get('chart', '')).strip().lower()
    if chart_id not in VALID_CHARTS:
        return jsonify({'error': f'Unknown chart id: {chart_id!r}'}), 400

    today     = datetime.utcnow().strftime('%Y-%m-%d')
    cache_key = f'{chart_id}_{today}'
    now       = time.time()

    # Serve from cache if valid
    cached = _analysis_cache.get(cache_key)
    if cached and now < cached['expires_at']:
        return jsonify({
            'analysis':    cached['analysis'],
            'generatedAt': cached['generated_at'],
            'cached':      True,
            'cacheAge':    round(now - cached['stored_at']),
        })

    # Load status.json for context
    try:
        with open(STATUS_JSON_PATH) as f:
            status = json.load(f)
    except Exception as e:
        return jsonify({'error': f'Market data unavailable: {e}'}), 503

    context     = CHART_CONTEXT_KEYS[chart_id](status)
    prompt_text = CHART_PROMPTS[chart_id].format(context_json=json.dumps(context, indent=2))

    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return jsonify({'error': 'Analysis service not configured (missing API key)'}), 503

    try:
        from anthropic import Anthropic as _Anthropic
        client = _Anthropic(api_key=api_key)
        msg = client.messages.create(
            model='claude-sonnet-4-20250514',
            max_tokens=600,
            temperature=0.3,
            messages=[{'role': 'user', 'content': prompt_text}],
        )
        analysis_text = msg.content[0].text.strip()
    except Exception as e:
        return jsonify({'error': f'Analysis generation failed: {e}'}), 503

    generated_at = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
    _analysis_cache[cache_key] = {
        'analysis':    analysis_text,
        'generated_at': generated_at,
        'stored_at':   now,
        'expires_at':  now + ANALYSIS_TTL,
    }
    return jsonify({
        'analysis':    analysis_text,
        'generatedAt': generated_at,
        'cached':      False,
        'cacheAge':    0,
    })



# ---------------------------------------------------------------------------
# QUARTERLY_JSON_PATH — data store for WGC demand data (written by /api/save-data)
# ---------------------------------------------------------------------------
QUARTERLY_JSON_PATH  = '/var/www/gold-tracker/data/quarterly.json'
WGC_DEMAND_JSON_PATH = '/var/www/gold-tracker/data/wgc_demand.json'


def parse_wgc_pdf(pdf_bytes):
    """Extract quarterly demand data from a WGC Gold Demand Trends PDF.

    Handles two PDF layouts:
    - Split-table: each data row is its own 1-row table (Q4/FY annual reports)
    - Classic: one multi-row table with header in row 0 (quarterly reports)

    Returns (rows, warnings).
    """
    import re

    def parse_col_header(text):
        """Return a 'YYYY-QN' key for quarter columns; None for everything else."""
        if not text:
            return None
        t = str(text).strip().replace('\n', ' ')
        # Skip change/pct/annual columns
        tl = t.lower()
        if any(x in tl for x in ['y/y', '%', 'change', 'annual', '△', '▲', '▼']):
            return None
        # Q4'24 / Q4'25 — ASCII or curly apostrophe
        m = re.match(r"Q([1-4])[\'\u2018\u2019](\d{2})$", t)
        if m:
            return f"20{m.group(2)}-Q{m.group(1)}"
        # Q4 2024
        m = re.match(r"Q([1-4])\s+(\d{4})$", t, re.I)
        if m:
            return f"{m.group(2)}-Q{m.group(1)}"
        # 2024 Q4
        m = re.match(r"(\d{4})\s+Q([1-4])$", t, re.I)
        if m:
            return f"{m.group(1)}-Q{m.group(2)}"
        return None

    # Row labels → internal field names
    # 'bars' and 'medals' are sub-items used to derive barCoin when no direct barCoin row
    ROW_MAP = {
        'jewellery consumption': 'jewelry',
        'jewellery':             'jewelry',
        'jewelry':               'jewelry',
        'bar & coin':            'barCoin',
        'bar and coin':          'barCoin',
        'physical bar':          'barCoin',
        'etf and similar':       'etfFlows',
        'etf':                   'etfFlows',
        'investment':            'investment',
        'bars':                  'bars',
        'medals imitation':      'medals',
        'medals':                'medals',
        'central bank':          'netPurchases',
        'net purchases':         'netPurchases',
        'lbma gold price':       'goldPrice',
        'lbma gold':             'goldPrice',
        'gold price':            'goldPrice',
        'average gold':          'goldPrice',
    }

    def map_row_label(label):
        lower = str(label or '').lower().strip()
        # Exact / prefix match first (avoids "etf and similar" → 'etf' misfire)
        for key, field in ROW_MAP.items():
            if lower == key or lower.startswith(key + ' ') or lower.startswith(key):
                return field
        # Substring fallback
        for key, field in ROW_MAP.items():
            if key in lower:
                return field
        return None

    def safe_float(val):
        s = re.sub(r'[^\d.\-]', '', str(val or '').replace(',', '').strip())
        try:
            return float(s) if s not in ('', '-') else None
        except ValueError:
            return None

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        all_page_tables = [page.extract_tables() or [] for page in pdf.pages]

    # ── Strategy 1: split-table layout (FY/Q4 annual reports) ────────────────
    # On one page: a header table (≥1 quarterly cols) + many single-row data tables
    best = {'col_map': {}, 'data': {}, 'score': 0}

    for page_tables in all_page_tables:
        # Find the column map from whichever table on this page has the most quarter cols
        col_map = {}
        for table in page_tables:
            if not table:
                continue
            candidate = {}
            for ci, cell in enumerate(table[0]):
                if ci == 0:
                    continue
                qkey = parse_col_header(str(cell or ''))
                if qkey:
                    candidate[ci] = qkey
            if len(candidate) > len(col_map):
                col_map = candidate

        if not col_map:
            continue

        min_col = max(col_map.keys()) + 1   # rows must be at least this wide

        data = {}  # field -> {quarter_key: value}
        for table in page_tables:
            for row in (table or []):
                if not row or not row[0]:
                    continue
                if len(row) < min_col:
                    continue
                field = map_row_label(str(row[0]))
                if not field:
                    continue
                for ci, qkey in col_map.items():
                    v = safe_float(row[ci] if ci < len(row) else None)
                    if v is not None:
                        data.setdefault(field, {})[qkey] = v

        score = sum(1 for f in ['jewelry', 'investment', 'bars', 'netPurchases', 'goldPrice']
                    if f in data)
        if score > best['score']:
            best = {'col_map': col_map, 'data': data, 'score': score}

    # ── Strategy 2: classic multi-row table layout (quarterly reports) ────────
    if best['score'] < 2:
        all_tables = [t for pt in all_page_tables for t in pt]
        # Find the first table that has "jewel" anywhere in col 0
        for table in all_tables:
            if not any(table[0]):
                continue
            # Check if this table or a nearby row has quarter headers
            col_map2 = {ci: parse_col_header(str(cell or ''))
                        for ci, cell in enumerate(table[0][1:], start=1)
                        if parse_col_header(str(cell or ''))}
            if len(col_map2) < 2:
                continue
            if not any('jewel' in str(r[0] or '').lower() for r in table[1:] if r):
                continue
            # Found it
            data2 = {}
            for row in table[1:]:
                if not row or not row[0]:
                    continue
                field = map_row_label(str(row[0]))
                if not field:
                    continue
                for ci, qkey in col_map2.items():
                    v = safe_float(row[ci] if ci < len(row) else None)
                    if v is not None:
                        data2.setdefault(field, {})[qkey] = v
            score2 = sum(1 for f in ['jewelry', 'etfFlows', 'barCoin', 'netPurchases'] if f in data2)
            if score2 > best['score']:
                best = {'col_map': col_map2, 'data': data2, 'score': score2}
            break

    col_map = best['col_map']
    data    = best['data']

    if not col_map:
        return [], ['No quarterly columns found in PDF']
    if not data:
        return [], ['Quarterly columns found but no data rows extracted']

    # ── Build output rows ─────────────────────────────────────────────────────
    quarters = sorted(set(col_map.values()))
    warnings = []
    rows = []

    for q in quarters:
        get = lambda f: data.get(f, {}).get(q)    # noqa: E731

        jewelry      = get('jewelry')
        barCoin      = get('barCoin')
        etf          = get('etfFlows')
        investment   = get('investment')
        bars         = get('bars')
        medals       = get('medals')
        netPurchases = get('netPurchases')
        goldPrice    = get('goldPrice')

        # Derive barCoin from sub-items when no direct "bar & coin" row
        if barCoin is None and bars is not None:
            barCoin = round(bars + (medals or 0), 1)
            if medals is None:
                warnings.append(f'{q}: bar & coin taken from "Bars" only (medals not found)')

        # Derive ETF from Investment − barCoin
        if etf is None and investment is not None and barCoin is not None:
            etf = round(investment - barCoin, 1)

        if jewelry is None or barCoin is None or netPurchases is None:
            continue

        rows.append({
            'quarter':      q,
            'jewelry':      jewelry,
            'barCoin':      barCoin,
            'etfFlows':     etf,
            'netPurchases': netPurchases,
            'goldPrice':    goldPrice,
        })

    if not rows:
        warnings.append('Data found but could not build complete rows '
                        '(need jewelry, barCoin, and netPurchases for each quarter).')
    elif any(r.get('barCoin') is not None and r.get('etfFlows') is not None
             and r['barCoin'] == r.get('bars') for r in rows):
        warnings.append('Note: bar & coin derived from Bars + Medals sub-items; '
                        'official coins not shown separately in this report format.')

    return rows, warnings


def parse_wgc_xlsx(xlsx_bytes):
    """
    Parse a WGC Gold Demand Trends XLSX file (GDT_Tables_Q*_EN.xlsx).

    Returns: (rows, warnings)
      rows     — list of dicts, one per quarter, keys:
                   quarter (str, e.g. '2025-Q4'),
                   central_banks (float, tonnes),
                   etf_flows (float, tonnes),
                   jewellery (float, tonnes),
                   bar_coin (float, tonnes),
                   total_demand (float, tonnes),
                   gold_price (float, USD/oz)
      warnings — list of strings (non-fatal issues)
    """
    import re
    import openpyxl

    warnings = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Cannot open XLSX: {e}")

    sheet_name = 'Gold Balance'
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}. "
            "Make sure this is a WGC Gold Demand Trends XLSX file."
        )

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    # Row 5 (index 4) contains headers including quarter labels
    header_row = all_rows[4]

    # Map column index -> normalized quarter string ('YYYY-QN')
    def normalize_quarter(label):
        m = re.match(r"Q(\d)'(\d{2})$", str(label).strip())
        if not m:
            return None
        qnum, yr2 = m.group(1), m.group(2)
        year = 2000 + int(yr2)
        return f"{year}-Q{qnum}"

    quarter_col_map = {}  # col_index -> 'YYYY-QN'
    for col_idx, val in enumerate(header_row):
        if val is None:
            continue
        normed = normalize_quarter(val)
        if normed:
            quarter_col_map[col_idx] = normed

    if not quarter_col_map:
        raise ValueError(
            "No quarterly columns found in row 5. "
            f"Sample row 5 values: {[v for v in header_row if v is not None][:10]}"
        )

    # Row definitions (0-indexed): label -> row_index
    ROW_MAP = {
        'central_banks': 24,   # Row 25
        'etf_flows':     23,   # Row 24
        'jewellery':     11,   # Row 12
        'bar_coin':      19,   # Row 20
        'total_demand':  10,   # Row 11
        'gold_price':    27,   # Row 28
    }

    def safe_float(val):
        if val is None or val == '' or (isinstance(val, str) and val.strip() in ('', '-', '▲', '▼')):
            return None
        try:
            return float(val)
        except (TypeError, ValueError):
            return None

    # Extract data row for each metric
    metric_data = {}
    for metric, row_idx in ROW_MAP.items():
        if row_idx >= len(all_rows):
            warnings.append(f"Row {row_idx + 1} not found for {metric}")
            metric_data[metric] = {}
            continue
        data_row = all_rows[row_idx]
        metric_data[metric] = {
            quarter_col_map[col_idx]: safe_float(data_row[col_idx])
            for col_idx in quarter_col_map
            if col_idx < len(data_row)
        }

    # Validate row labels match expectations
    label_checks = {
        24: 'Central Bank',
        23: 'ETF',
        11: 'Jewellery',
        19: 'Bar and Coin',
        10: 'Total Demand',
        27: 'LBMA',
    }
    for row_idx, expected_substr in label_checks.items():
        label = str(all_rows[row_idx][1] or '').strip()
        if expected_substr.lower() not in label.lower():
            warnings.append(
                f"Row {row_idx + 1} label mismatch: expected '{expected_substr}', "
                f"got '{label}' — data may be misaligned"
            )

    # Build output rows, sorted by quarter
    all_quarters = sorted(set(metric_data['central_banks'].keys()))

    # Only include quarters where we have at least CB demand data
    rows_out = []
    for q in all_quarters:
        cb = metric_data['central_banks'].get(q)
        if cb is None:
            continue
        rows_out.append({
            'quarter':       q,
            'central_banks': cb,
            'etf_flows':     metric_data['etf_flows'].get(q),
            'jewellery':     metric_data['jewellery'].get(q),
            'bar_coin':      metric_data['bar_coin'].get(q),
            'total_demand':  metric_data['total_demand'].get(q),
            'gold_price':    metric_data['gold_price'].get(q),
        })

    if not rows_out:
        raise ValueError("Parsed 0 data rows from XLSX. Check file format.")

    return rows_out, warnings


@app.route('/api/upload-pdf', methods=['POST'])
def upload_pdf():
    """Accept a WGC Gold Demand Trends PDF, extract quarterly data, return rows as JSON."""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        f = request.files['file']
        if not f.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'File must be a PDF'}), 400
        pdf_bytes = f.read()
        rows, warnings = parse_wgc_pdf(pdf_bytes)
        if not rows:
            return jsonify({'error': 'Could not extract quarterly data from PDF. '
                                     'Make sure this is a WGC Gold Demand Trends report.'}), 422
        return jsonify({'rows': rows, 'warnings': warnings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/upload-xlsx', methods=['POST'])
def upload_xlsx():
    """Accept a WGC Gold Demand Trends XLSX, parse it, return rows as JSON for confirmation."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    xlsx_bytes = request.files['file'].read()
    try:
        rows, warnings = parse_wgc_xlsx(xlsx_bytes)
    except (ValueError, Exception) as e:
        return jsonify({'error': str(e),
                        'hint': 'Make sure this is a WGC Gold Demand Trends XLSX file '
                                '(e.g. GDT_Tables_Q425_EN.xlsx)'}), 422
    return jsonify({'rows': rows, 'warnings': warnings, 'count': len(rows)})


@app.route('/api/save-xlsx-data', methods=['POST'])
def save_xlsx_data():
    """Persist confirmed XLSX rows (list format with central_banks key) to wgc_demand.json."""
    try:
        rows = request.get_json(force=True)
        if not isinstance(rows, list) or not rows:
            return jsonify({'error': 'Expected a non-empty JSON array of rows'}), 400
        required = {'quarter', 'central_banks'}
        missing = [i for i, r in enumerate(rows) if not required.issubset(r.keys())]
        if missing:
            return jsonify({'error': f'Rows at indices {missing[:5]} missing required keys'}), 400
        os.makedirs(os.path.dirname(WGC_DEMAND_JSON_PATH), exist_ok=True)
        with open(WGC_DEMAND_JSON_PATH, 'w') as f:
            json.dump(rows, f, indent=2)
        write_status_json()   # regenerate scorecard with new CB demand signal
        return jsonify({'ok': True, 'rows': len(rows),
                        'latestQuarter': max(r['quarter'] for r in rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/upload-pdf-debug', methods=['POST'])
def upload_pdf_debug():
    """Debug endpoint: return raw pdfplumber table extraction for diagnosis."""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        f = request.files['file']
        pdf_bytes = f.read()
        result = {'pages': [], 'all_tables': []}
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            result['num_pages'] = len(pdf.pages)
            for pi, page in enumerate(pdf.pages):
                tables = page.extract_tables() or []
                page_info = {'page': pi + 1, 'num_tables': len(tables), 'tables': []}
                for ti, table in enumerate(tables):
                    page_info['tables'].append({
                        'table_index': ti,
                        'rows': len(table),
                        'cols': len(table[0]) if table else 0,
                        'all_rows': table,
                    })
                result['pages'].append(page_info)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/save-data', methods=['POST'])
def save_data():
    """Accept a full quarterlyData JSON payload from the browser and persist it to disk.
    Called after the user confirms a CSV upload on the frontend."""
    try:
        data = request.get_json(force=True)
        required = ['netPurchases', 'holdings', 'goldPrice', 'jewelry', 'barCoin', 'etfFlows']
        if not data or not all(k in data for k in required):
            return jsonify({'error': 'Missing required data keys: ' + str(required)}), 400
        # Sanity: at least one quarter in each series
        if any(not data[k] for k in required):
            return jsonify({'error': 'One or more data series is empty'}), 400
        os.makedirs(os.path.dirname(QUARTERLY_JSON_PATH), exist_ok=True)
        with open(QUARTERLY_JSON_PATH, 'w') as f:
            json.dump(data, f, indent=2)
        n = len(data['jewelry'])
        return jsonify({'ok': True, 'quarters': n,
                        'latestQuarter': sorted(data['jewelry'].keys())[-1]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5000, debug=False)
